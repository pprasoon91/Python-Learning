import os
import openpyxl
from openpyxl import Workbook
from datetime import datetime

def list_files_to_excel(folder_path, output_excel="file_list_with_date.xlsx"):
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "File List"
    
    # Set headers
    ws['A1'] = 'Filename'
    ws['B1'] = 'File Path'
    ws['C1'] = 'Modified Date'
    
    row = 2  # Start from second row (after headers)

    for root, _, files in os.walk(folder_path):
        for file in files:
            full_path = os.path.join(root, file)
            modified_time = os.path.getmtime(full_path)
            formatted_time = datetime.fromtimestamp(modified_time).strftime("%Y-%m-%d %H:%M:%S")

            ws.cell(row=row, column=1, value=file)
            ws.cell(row=row, column=2, value=full_path)
            ws.cell(row=row, column=3, value=formatted_time)
            row += 1

    # Save the Excel file
    wb.save(output_excel)
    print(f"Excel file saved as '{output_excel}' with {row - 2} entries.")

if __name__ == "__main__":
    folder_path = input("Enter the full path of the folder: ").strip()
    if os.path.isdir(folder_path):
        list_files_to_excel(folder_path)
    else:
        print("Invalid folder path. Please try again.")
